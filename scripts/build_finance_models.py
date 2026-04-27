#!/usr/bin/env python3
"""build_finance_models.py — Generate the Modulor finance spreadsheets.

Produces two .xlsx files for the Drive data room (`01_Finances/`):

  Modulor-Use-of-Funds-Model-v1.xlsx
    - Allocation sheet ($3M split per CLAUDE.md priorities)
    - Monthly Deployment sheet (categories × 30 months, w/ totals)
    - Assumptions sheet (drivers, edit + rerun)

  Modulor-Burn-Forecast.xlsx
    - Pre-funding Burn (Apr 2026 → close)
    - Post-funding Burn (close → Mo 30, w/ running cash + runway)
    - Assumptions sheet

Edit assumptions at the top of this file, then re-run.

Usage:
    python scripts/build_finance_models.py
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "build" / "drive-upload" / "01_Finances"

# ─────────────────────────────────────────────────────────────────────────────
# Assumptions (edit + rerun)
# ─────────────────────────────────────────────────────────────────────────────

RAISE_AMOUNT = 3_000_000
HORIZON_MONTHS = 30
TARGET_CLOSE = date(2026, 9, 1)         # post-funding Month 1 starts here
PRE_FUNDING_START = date(2026, 4, 1)    # Apr 2026 (formation month)

# Allocation per CLAUDE.md priorities. Must sum to RAISE_AMOUNT.
ALLOCATION = {
    "Hardware prototyping": 1_200_000,   # 40%
    "Engineering":            900_000,   # 30%
    "Go-to-market":           450_000,   # 15%
    "IP / entity / data":     300_000,   # 10%
    "Legal":                  150_000,   #  5%
}

# Phasing weights — one weight per month over HORIZON_MONTHS. Each row is
# normalized so weights × allocation / sum(weights) = monthly spend. Edit to
# reshape the curve.
PHASING = {
    # Hardware: heavy front-load for prototype iterations Mo 1-9, taper after
    "Hardware prototyping": [
        1.5, 1.6, 1.7, 1.7, 1.6, 1.5, 1.4, 1.2, 1.0,   # Mo 1-9   peak build
        0.7, 0.6, 0.6, 0.5, 0.5, 0.4, 0.4, 0.4, 0.3,   # Mo 10-18 taper
        0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3,  # Mo 19-30 maint
    ],
    # Engineering: contractors Mo 1-5, FT hire Mo 6+, peak Mo 12-24
    "Engineering": [
        0.6, 0.7, 0.8, 0.9, 0.9, 1.0, 1.1, 1.1, 1.1,   # Mo 1-9   ramp
        1.1, 1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.1, 1.1,   # Mo 10-18 peak
        1.1, 1.1, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0,  # Mo 19-30 steady
    ],
    # GTM: light Mo 1-12 (research + pilot lead-gen), heavy Mo 13-30 (pilots, BD hire)
    "Go-to-market": [
        0.4, 0.4, 0.4, 0.4, 0.5, 0.5, 0.6, 0.6, 0.7,   # Mo 1-9   research
        0.7, 0.8, 0.9, 1.0, 1.2, 1.3, 1.4, 1.5, 1.5,   # Mo 10-18 ramp
        1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.4, 1.4, 1.4, 1.4, 1.4, 1.4,  # Mo 19-30 sustain
    ],
    # IP: lumpy — non-prov conversion Mo 8 ($10K), trademark Mo 1-2, ongoing data
    "IP / entity / data": [
        1.5, 1.5, 0.6, 0.6, 0.6, 0.7, 0.8, 3.5, 1.0,   # Mo 1-9   trademark + non-prov spike Mo 8
        1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0,   # Mo 10-18 data infra
        1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0.9, 0.9, 0.8, 0.8, 0.8, 0.8,  # Mo 19-30 sustain
    ],
    # Legal: front-loaded — RSPA + 83(b) + SAFE redlines + counsel retainer Mo 1-6
    "Legal": [
        2.5, 2.0, 1.8, 1.5, 1.3, 1.2, 1.0, 0.9, 0.9,   # Mo 1-9   formation + raise close
        0.8, 0.8, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7,   # Mo 10-18 ongoing
        0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7,  # Mo 19-30 quarterly reviews
    ],
}

# Pre-funding burn: from formation (Apr 2026) through target close month
PRE_FUNDING_LINES = [
    # (description, amount, month_offset_from_PRE_FUNDING_START, status)
    ("Stripe Atlas formation",                       500.00, 0, "✅ Paid (founder Amex)"),
    ("Stable virtual address (annual)",              588.00, 0, "✅ Paid (founder Amex)"),
    ("Google Workspace seat (Mitchell, monthly)",     14.00, 0, "Recurring"),
    ("Google Workspace seat (Mitchell, monthly)",     14.00, 1, "Recurring"),
    ("Google Workspace seat (Mitchell, monthly)",     14.00, 2, "Recurring"),
    ("Google Workspace seat (Mitchell, monthly)",     14.00, 3, "Recurring"),
    ("Google Workspace seat (Mitchell, monthly)",     14.00, 4, "Recurring"),
    ("Google Workspace seat (Luke, monthly)",         14.00, 1, "Recurring (Luke active May+)"),
    ("Google Workspace seat (Luke, monthly)",         14.00, 2, "Recurring"),
    ("Google Workspace seat (Luke, monthly)",         14.00, 3, "Recurring"),
    ("Google Workspace seat (Luke, monthly)",         14.00, 4, "Recurring"),
    ("Slack Pro (2 seats, monthly)",                  17.00, 1, "Recurring (Apr 27 setup)"),
    ("Slack Pro (2 seats, monthly)",                  17.00, 2, "Recurring"),
    ("Slack Pro (2 seats, monthly)",                  17.00, 3, "Recurring"),
    ("Slack Pro (2 seats, monthly)",                  17.00, 4, "Recurring"),
    ("Outside counsel retainer (Atlas Plus partner)",2_500.00, 1, "Pending — Apr 27 AM submit"),
    ("RSPA + 83(b) execution package (Luke)",         750.00, 1, "Pending — same package"),
    ("Patent assignment Mitchell → Modulor",          500.00, 1, "Pending — same package"),
    ("USPTO trademark filing (3 classes, TEAS Plus)",1_050.00, 2, "Self-file (no counsel)"),
    ("D&O insurance, Year 1 (mid-quote)",           2_250.00, 3, "Required pre-first-check"),
    ("Mutual NDA (Adam, vendors) — counsel time",     250.00, 1, "Pending"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Styling
# ─────────────────────────────────────────────────────────────────────────────

H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
H2 = Font(name="Calibri", size=12, bold=True)
BODY = Font(name="Calibri", size=11)
BODY_BOLD = Font(name="Calibri", size=11, bold=True)
MUTED = Font(name="Calibri", size=10, color="666666", italic=True)

FILL_HEADER = PatternFill("solid", fgColor="2E4C3D")    # Modulor green
FILL_SUBHEAD = PatternFill("solid", fgColor="DCE7DF")
FILL_TOTAL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(border_style="thin", color="CCCCCC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

USD_FMT = '_($* #,##0_);[Red]_($* (#,##0);_($* "-"??_);_(@_)'
PCT_FMT = "0.0%"


def add_month(d: date, n: int) -> date:
    """Return d + n months (1st of month)."""
    m = d.month - 1 + n
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, 1)


def style_header_row(ws, row: int, last_col: int, label_col: int = 1) -> None:
    for c in range(label_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H1
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BOX


# ─────────────────────────────────────────────────────────────────────────────
# Use-of-Funds workbook
# ─────────────────────────────────────────────────────────────────────────────

def build_use_of_funds() -> Path:
    wb = Workbook()

    # ─── Sheet 1: Allocation ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Allocation"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Modulor, Inc. — $3M Use of Funds (Model v1)"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="2E4C3D")
    ws.merge_cells("A1:D1")

    ws["A2"] = (
        f"Target raise: ${RAISE_AMOUNT:,.0f} · target close {TARGET_CLOSE:%b %Y} · "
        f"{HORIZON_MONTHS}-month runway · ${RAISE_AMOUNT/HORIZON_MONTHS:,.0f}/mo avg burn"
    )
    ws["A2"].font = MUTED
    ws.merge_cells("A2:D2")

    headers = ["Category", "Amount", "% of raise", "Drivers"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 4)

    drivers = {
        "Hardware prototyping": "Engineering firm fees (Mo 1–9), 3–5 prototype iterations, BOM, testing rigs, pilot-site units",
        "Engineering":          "Contracted firmware/ML (Mo 1–6), FT engineering hire #1 (Mo 6+), sensor integration, app dev",
        "Go-to-market":         "Pilot-site outfitting, sales materials, conference presence, first BD hire (Mo 12+)",
        "IP / entity / data":   "Patent non-prov conversion (Mo 8), trademark, data licensing infra, additional provisionals",
        "Legal":                "Outside counsel retainer, term-sheet redlines, board/governance, ongoing IP support",
    }

    row = 5
    for cat, amount in ALLOCATION.items():
        ws.cell(row=row, column=1, value=cat).font = BODY_BOLD
        c2 = ws.cell(row=row, column=2, value=amount)
        c2.number_format = USD_FMT
        c2.alignment = RIGHT
        c3 = ws.cell(row=row, column=3, value=amount / RAISE_AMOUNT)
        c3.number_format = PCT_FMT
        c3.alignment = RIGHT
        ws.cell(row=row, column=4, value=drivers[cat]).alignment = LEFT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BOX
        row += 1

    # Total
    ws.cell(row=row, column=1, value="Total").font = BODY_BOLD
    t2 = ws.cell(row=row, column=2, value=f"=SUM(B5:B{row-1})")
    t2.font = BODY_BOLD
    t2.number_format = USD_FMT
    t2.fill = FILL_TOTAL
    t3 = ws.cell(row=row, column=3, value=f"=SUM(C5:C{row-1})")
    t3.font = BODY_BOLD
    t3.number_format = PCT_FMT
    t3.fill = FILL_TOTAL
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOX
        if c != 4:
            ws.cell(row=row, column=c).fill = FILL_TOTAL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 80

    # ─── Sheet 2: Monthly Deployment ───────────────────────────────────────
    md = wb.create_sheet("Monthly Deployment")
    md.sheet_view.showGridLines = False
    md.freeze_panes = "B3"

    md["A1"] = f"Monthly deployment — Mo 1 = {TARGET_CLOSE:%b %Y}, $ to nearest $1"
    md["A1"].font = H2
    md.merge_cells(start_row=1, start_column=1, end_row=1, end_column=HORIZON_MONTHS + 2)

    # Header row
    md.cell(row=2, column=1, value="Category")
    for m in range(HORIZON_MONTHS):
        d = add_month(TARGET_CLOSE, m)
        md.cell(row=2, column=2 + m, value=f"Mo {m+1}\n{d:%b %y}")
    md.cell(row=2, column=2 + HORIZON_MONTHS, value="Total")
    style_header_row(md, 2, 2 + HORIZON_MONTHS)

    # Compute monthly deployment per category (normalize phasing × allocation)
    monthly = {}
    for cat, alloc in ALLOCATION.items():
        weights = PHASING[cat]
        assert len(weights) == HORIZON_MONTHS, f"phasing length mismatch: {cat}"
        total_w = sum(weights)
        monthly[cat] = [round(w * alloc / total_w) for w in weights]
        # Adjust last month to absorb rounding drift
        drift = alloc - sum(monthly[cat])
        monthly[cat][-1] += drift

    row = 3
    for cat in ALLOCATION:
        md.cell(row=row, column=1, value=cat).font = BODY_BOLD
        for m in range(HORIZON_MONTHS):
            cell = md.cell(row=row, column=2 + m, value=monthly[cat][m])
            cell.number_format = USD_FMT
            cell.alignment = RIGHT
        total_cell = md.cell(row=row, column=2 + HORIZON_MONTHS,
                             value=f"=SUM(B{row}:{get_column_letter(1+HORIZON_MONTHS)}{row})")
        total_cell.number_format = USD_FMT
        total_cell.font = BODY_BOLD
        total_cell.fill = FILL_TOTAL
        row += 1

    # Monthly total row
    md.cell(row=row, column=1, value="Monthly burn").font = BODY_BOLD
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        cell = md.cell(row=row, column=2 + m,
                       value=f"=SUM({col}3:{col}{row-1})")
        cell.number_format = USD_FMT
        cell.font = BODY_BOLD
        cell.fill = FILL_SUBHEAD
    grand = md.cell(row=row, column=2 + HORIZON_MONTHS,
                    value=f"=SUM(B{row}:{get_column_letter(1+HORIZON_MONTHS)}{row})")
    grand.number_format = USD_FMT
    grand.font = H2
    grand.fill = FILL_SUBHEAD
    row += 1

    # Cumulative spend row
    md.cell(row=row, column=1, value="Cumulative spend").font = MUTED
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        if m == 0:
            cell = md.cell(row=row, column=2 + m, value=f"={col}{row-1}")
        else:
            prev = get_column_letter(1 + m)
            cell = md.cell(row=row, column=2 + m,
                           value=f"={prev}{row}+{col}{row-1}")
        cell.number_format = USD_FMT
    row += 1

    # Cash remaining row
    md.cell(row=row, column=1, value="Cash remaining").font = MUTED
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        cell = md.cell(row=row, column=2 + m,
                       value=f"={RAISE_AMOUNT}-{col}{row-1}")
        cell.number_format = USD_FMT
    row += 1

    md.column_dimensions["A"].width = 26
    for m in range(HORIZON_MONTHS):
        md.column_dimensions[get_column_letter(2 + m)].width = 11
    md.column_dimensions[get_column_letter(2 + HORIZON_MONTHS)].width = 14
    md.row_dimensions[2].height = 32

    # ─── Sheet 3: Assumptions ──────────────────────────────────────────────
    a = wb.create_sheet("Assumptions")
    a.sheet_view.showGridLines = False
    a["A1"] = "Assumptions (edit + rerun script to update)"
    a["A1"].font = H2

    rows = [
        ("Raise amount",         f"${RAISE_AMOUNT:,.0f}"),
        ("Target close",         f"{TARGET_CLOSE:%B %Y}"),
        ("Horizon",              f"{HORIZON_MONTHS} months"),
        ("Avg monthly burn",     f"${RAISE_AMOUNT/HORIZON_MONTHS:,.0f}"),
        ("Allocation source",    "CLAUDE.md priorities — 40/30/15/10/5"),
        ("Hardware curve",       "Front-loaded Mo 1–9 (prototype iterations); taper Mo 10–30 (pilot units, refresh)"),
        ("Engineering curve",    "Contractor ramp Mo 1–5; FT hire Mo 6+; peak Mo 12–24; steady Mo 25–30"),
        ("GTM curve",            "Light Mo 1–12 (research, pilot lead-gen); heavy Mo 13–30 (pilots, BD hire)"),
        ("IP curve",             "Trademark Mo 1–2; non-prov conversion spike Mo 8 (~$10K); steady data infra Mo 9–30"),
        ("Legal curve",          "Front-loaded Mo 1–6 (RSPA, 83(b), SAFE redlines, retainer); quarterly reviews after"),
        ("Rounding",             "Last-month rebalanced so category totals tie to allocation exactly"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        a.cell(row=i, column=1, value=k).font = BODY_BOLD
        a.cell(row=i, column=2, value=v).font = BODY
    a.column_dimensions["A"].width = 22
    a.column_dimensions["B"].width = 90

    out = OUT_DIR / "Modulor-Use-of-Funds-Model-v1.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Burn Forecast workbook
# ─────────────────────────────────────────────────────────────────────────────

def build_burn_forecast() -> Path:
    wb = Workbook()

    # ─── Sheet 1: Pre-funding Burn ─────────────────────────────────────────
    pre = wb.active
    pre.title = "Pre-funding Burn"
    pre.sheet_view.showGridLines = False

    pre["A1"] = (
        f"Pre-funding burn — {PRE_FUNDING_START:%b %Y} → {TARGET_CLOSE:%b %Y} "
        f"(target close)"
    )
    pre["A1"].font = Font(name="Calibri", size=16, bold=True, color="2E4C3D")
    pre.merge_cells("A1:E1")

    pre["A2"] = "Founder-funded; reimburse from Modulor account once Mercury opens."
    pre["A2"].font = MUTED
    pre.merge_cells("A2:E2")

    headers = ["Month", "Date", "Item", "Amount", "Status"]
    for i, h in enumerate(headers, 1):
        pre.cell(row=4, column=i, value=h)
    style_header_row(pre, 4, len(headers))

    row = 5
    for desc, amount, offset, status in PRE_FUNDING_LINES:
        d = add_month(PRE_FUNDING_START, offset)
        pre.cell(row=row, column=1, value=f"M{offset+1}")
        pre.cell(row=row, column=2, value=d.strftime("%b %Y"))
        pre.cell(row=row, column=3, value=desc)
        c4 = pre.cell(row=row, column=4, value=amount)
        c4.number_format = USD_FMT
        c4.alignment = RIGHT
        pre.cell(row=row, column=5, value=status)
        for c in range(1, 6):
            pre.cell(row=row, column=c).border = BOX
        row += 1

    # Total
    pre.cell(row=row, column=3, value="Total pre-funding spend").font = BODY_BOLD
    t = pre.cell(row=row, column=4, value=f"=SUM(D5:D{row-1})")
    t.font = H2
    t.number_format = USD_FMT
    t.fill = FILL_TOTAL
    for c in range(1, 6):
        pre.cell(row=row, column=c).border = BOX

    pre.column_dimensions["A"].width = 8
    pre.column_dimensions["B"].width = 11
    pre.column_dimensions["C"].width = 48
    pre.column_dimensions["D"].width = 14
    pre.column_dimensions["E"].width = 32

    # ─── Sheet 2: Post-funding Burn ────────────────────────────────────────
    post = wb.create_sheet("Post-funding Burn")
    post.sheet_view.showGridLines = False
    post.freeze_panes = "B3"

    post["A1"] = (
        f"Post-funding burn — Mo 1 = {TARGET_CLOSE:%b %Y} · "
        f"raise ${RAISE_AMOUNT:,.0f} · {HORIZON_MONTHS} months"
    )
    post["A1"].font = H2
    post.merge_cells(start_row=1, start_column=1, end_row=1, end_column=HORIZON_MONTHS + 2)

    post.cell(row=2, column=1, value="Line")
    for m in range(HORIZON_MONTHS):
        d = add_month(TARGET_CLOSE, m)
        post.cell(row=2, column=2 + m, value=f"Mo {m+1}\n{d:%b %y}")
    post.cell(row=2, column=2 + HORIZON_MONTHS, value="Total")
    style_header_row(post, 2, 2 + HORIZON_MONTHS)

    # Re-compute monthly per category (same as use-of-funds sheet)
    monthly = {}
    for cat, alloc in ALLOCATION.items():
        weights = PHASING[cat]
        total_w = sum(weights)
        monthly[cat] = [round(w * alloc / total_w) for w in weights]
        drift = alloc - sum(monthly[cat])
        monthly[cat][-1] += drift

    row = 3
    for cat in ALLOCATION:
        post.cell(row=row, column=1, value=cat).font = BODY
        for m in range(HORIZON_MONTHS):
            cell = post.cell(row=row, column=2 + m, value=monthly[cat][m])
            cell.number_format = USD_FMT
        tcell = post.cell(row=row, column=2 + HORIZON_MONTHS,
                          value=f"=SUM(B{row}:{get_column_letter(1+HORIZON_MONTHS)}{row})")
        tcell.number_format = USD_FMT
        tcell.font = BODY_BOLD
        tcell.fill = FILL_TOTAL
        row += 1

    monthly_burn_row = row
    post.cell(row=row, column=1, value="Monthly burn").font = BODY_BOLD
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        cell = post.cell(row=row, column=2 + m,
                         value=f"=SUM({col}3:{col}{row-1})")
        cell.number_format = USD_FMT
        cell.font = BODY_BOLD
        cell.fill = FILL_SUBHEAD
    g = post.cell(row=row, column=2 + HORIZON_MONTHS,
                  value=f"=SUM(B{row}:{get_column_letter(1+HORIZON_MONTHS)}{row})")
    g.number_format = USD_FMT
    g.font = H2
    g.fill = FILL_SUBHEAD
    row += 1

    # Cash in row (raise lands Mo 1)
    cash_in_row = row
    post.cell(row=row, column=1, value="Cash in").font = MUTED
    for m in range(HORIZON_MONTHS):
        cell = post.cell(row=row, column=2 + m, value=RAISE_AMOUNT if m == 0 else 0)
        cell.number_format = USD_FMT
    row += 1

    # Running cash
    running_row = row
    post.cell(row=row, column=1, value="Running cash (end of month)").font = BODY_BOLD
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        if m == 0:
            cell = post.cell(row=row, column=2 + m,
                             value=f"={col}{cash_in_row}-{col}{monthly_burn_row}")
        else:
            prev = get_column_letter(1 + m)
            cell = post.cell(row=row, column=2 + m,
                             value=f"={prev}{row}+{col}{cash_in_row}-{col}{monthly_burn_row}")
        cell.number_format = USD_FMT
        cell.font = BODY_BOLD
    row += 1

    # Months of runway remaining (running cash / 3-month avg forward burn)
    runway_row = row
    post.cell(row=row, column=1, value="Runway months remaining").font = MUTED
    for m in range(HORIZON_MONTHS):
        col = get_column_letter(2 + m)
        # Use trailing 3-month avg burn as denominator (or current month if early)
        if m < 2:
            denom = f"{col}{monthly_burn_row}"
        else:
            prev2 = get_column_letter(m)        # 2 cols back
            prev1 = get_column_letter(1 + m)    # 1 col back
            denom = f"AVERAGE({prev2}{monthly_burn_row}:{col}{monthly_burn_row})"
        cell = post.cell(row=row, column=2 + m,
                         value=f"=IFERROR({col}{running_row}/{denom},0)")
        cell.number_format = "0.0"
    row += 1

    post.column_dimensions["A"].width = 30
    for m in range(HORIZON_MONTHS):
        post.column_dimensions[get_column_letter(2 + m)].width = 11
    post.column_dimensions[get_column_letter(2 + HORIZON_MONTHS)].width = 14
    post.row_dimensions[2].height = 32

    # ─── Sheet 3: Assumptions ──────────────────────────────────────────────
    a = wb.create_sheet("Assumptions")
    a.sheet_view.showGridLines = False
    a["A1"] = "Burn forecast assumptions"
    a["A1"].font = H2

    pre_window_months = (
        (TARGET_CLOSE.year - PRE_FUNDING_START.year) * 12
        + TARGET_CLOSE.month
        - PRE_FUNDING_START.month
    )
    pre_window = (
        f"{PRE_FUNDING_START:%b %Y} → {TARGET_CLOSE:%b %Y} ({pre_window_months} months)"
    )

    rows = [
        ("Pre-funding window",  pre_window),
        ("Pre-funding burn",    "Founder-funded one-shots + recurring Workspace + Slack; "
                                "reimbursed from Modulor account post-Mercury open."),
        ("Closing month",       f"{TARGET_CLOSE:%B %Y} (target). Cash lands Mo 1 of post-funding."),
        ("Raise amount",        f"${RAISE_AMOUNT:,.0f}"),
        ("Horizon",             f"{HORIZON_MONTHS} months post-funding"),
        ("Allocation",          "Per CLAUDE.md priorities (40/30/15/10/5)"),
        ("Phasing",             "See PHASING dict in scripts/build_finance_models.py — "
                                "edit + rerun to reshape"),
        ("Runway formula",      "Running cash / trailing 3-month avg burn (cell of current month)"),
        ("Refresh cadence",     "Update monthly on the 1st alongside Mercury reconciliation"),
    ]

    for i, (k, v) in enumerate(rows, 3):
        a.cell(row=i, column=1, value=k).font = BODY_BOLD
        a.cell(row=i, column=2, value=v).font = BODY
    a.column_dimensions["A"].width = 22
    a.column_dimensions["B"].width = 100

    out = OUT_DIR / "Modulor-Burn-Forecast.xlsx"
    wb.save(str(out))
    return out


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    uof = build_use_of_funds()
    burn = build_burn_forecast()
    print(f"  ✓ {uof.relative_to(REPO_ROOT)}")
    print(f"  ✓ {burn.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
